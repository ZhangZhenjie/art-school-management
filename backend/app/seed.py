"""Seed script — initial data import.

Run: `python -m app.seed` from inside backend/.

Stages:
1. schedules: insert the 32 fixed classes from spec §3.1.
2. users: insert Lucas (teacher) and Y (admin) with bcrypt-hashed passwords.
3. students + packages: TODO M1 finalize — read from SEED_EXCEL_PATH if set.
"""
from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

from .config import settings
from .database import Base, SessionLocal, engine
from .models.package import Package
from .models.schedule import Schedule
from .models.student import Student
from .models.user import User
from .routers.auth import hash_password

# Excel epoch: serial 0 == 1899-12-30 (handles 1900 leap-year quirk).
_EXCEL_EPOCH = datetime(1899, 12, 30)


def _excel_serial_to_date(serial: float | int | None) -> date | None:
    if serial is None:
        return None
    return (_EXCEL_EPOCH + timedelta(days=float(serial))).date()


def _clean_optional(v):
    """Normalize Excel cell to nullable string. NA / blank -> None."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.upper() == "NA":
        return None
    return s


SCHEDULES: list[tuple[str, str, int, str, str]] = [
    ("2C", "兴趣班", 2, "16:30", "18:30"),
    ("2F", "专业班", 2, "16:30", "18:30"),
    ("3C1", "兴趣班", 3, "15:00", "17:00"),
    ("3F1", "专业班", 3, "15:00", "17:00"),
    ("3K", "小小艺术家", 3, "15:30", "17:30"),
    ("3C2", "兴趣班", 3, "16:30", "18:30"),
    ("3F2", "专业班", 3, "16:30", "18:30"),
    ("4K", "小小艺术家", 4, "15:30", "17:00"),
    ("4C1", "兴趣班", 4, "15:30", "17:30"),
    ("4C2", "兴趣班", 4, "16:00", "18:00"),
    ("4F1", "专业班", 4, "16:00", "18:00"),
    ("4F2", "专业班", 4, "19:00", "21:00"),
    ("5A", "升学班", 5, "15:00", "17:00"),
    ("5H", "手工班", 5, "15:00", "17:00"),
    ("6C1", "兴趣班", 6, "09:00", "11:00"),
    ("6F1", "专业班", 6, "10:00", "12:00"),
    ("6K1", "小小艺术家", 6, "11:00", "12:30"),
    ("6C2", "兴趣班", 6, "11:00", "13:00"),
    ("6F2", "专业班", 6, "13:30", "16:00"),
    ("6C3", "兴趣班", 6, "14:00", "16:00"),
    ("6K2", "小小艺术家", 6, "16:00", "17:30"),
    ("6F3", "专业班", 6, "16:00", "18:00"),
    ("6C4", "兴趣班", 6, "16:00", "18:00"),
    ("7K1", "小小艺术家", 7, "10:30", "12:00"),
    ("7C1", "兴趣班", 7, "10:00", "12:00"),
    ("7F1", "专业班", 7, "10:00", "12:30"),
    ("7C2", "兴趣班", 7, "13:30", "15:30"),
    ("7F2", "专业班", 7, "13:30", "15:30"),
    ("7F3", "专业班", 7, "15:30", "18:00"),
    ("7K2", "小小艺术家", 7, "16:00", "17:30"),
    ("7C3", "兴趣班", 7, "16:00", "18:00"),
    ("7B", "成人班", 7, "16:00", "18:00"),
]

# spec §2 — hardcoded credentials.
INITIAL_USERS = [
    ("Lucassss", "and2026_", "teacher"),
    ("XY_", "and2026!", "admin"),
]


def seed_schedules(db) -> None:
    if db.query(Schedule).count() > 0:
        return
    for sid, name, weekday, start, end in SCHEDULES:
        db.add(Schedule(id=sid, name=name, weekday=weekday, start_time=start, end_time=end))
    db.commit()


def seed_users(db) -> None:
    if db.query(User).count() > 0:
        return
    for username, password, role in INITIAL_USERS:
        db.add(User(username=username, password_hash=hash_password(password), role=role))
    db.commit()


def seed_students_and_packages(db) -> None:
    """Import students + first packages from Excel at SEED_EXCEL_PATH.

    Sheets used (spec §3.2 / §3.3 / §5):
      - `student`:        id | name | current_schedule_id | email | birthday(serial) | parent_name | phone
      - `student_package`: id | student_id | student_name | num_classes | unit_price | start_date(serial) | end_date(serial) | is_negative | used_classes

    Per spec §5:
      - purchased_classes = total_classes = remaining_classes (initial simplification — gifted=0)
      - start_date = 2025-07-01, end_date = 2026-12-01
      - schedule_id "1O" preserved as soft reference (no FK)
      - birthday Excel serial -> ISO date
    """
    if db.query(Student).count() > 0:
        return
    if not settings.SEED_EXCEL_PATH:
        print("[seed] SEED_EXCEL_PATH not set, skipping students/packages.")
        return

    path = Path(settings.SEED_EXCEL_PATH)
    if not path.is_absolute():
        path = (Path(__file__).resolve().parent.parent / path).resolve()
    if not path.exists():
        print(f"[seed] Excel not found at {path}, skipping students/packages.")
        return

    wb = openpyxl.load_workbook(path, data_only=True)
    student_rows = list(wb["student"].iter_rows(min_row=2, values_only=True))
    package_rows = list(wb["student_package"].iter_rows(min_row=2, values_only=True))

    # Build package lookup by trimmed student_name (col index 2).
    pkg_by_name: dict[str, tuple] = {}
    for row in package_rows:
        if not row or row[2] is None:
            continue
        pkg_by_name[str(row[2]).strip()] = row

    # spec §5 fixed import dates.
    start_date = date(2025, 7, 1)
    end_date = date(2026, 12, 1)

    students_added = 0
    packages_added = 0
    missing_pkg: list[str] = []

    for row in student_rows:
        if not row or row[1] is None:
            continue
        name = str(row[1]).strip()
        schedule_id = _clean_optional(row[2])
        email = _clean_optional(row[3])
        birthday = _excel_serial_to_date(row[4]) if row[4] is not None else None
        parent_name = _clean_optional(row[5])
        phone = _clean_optional(row[6])

        student = Student(
            name=name,
            birthday=birthday,
            email=email,
            parent_name=parent_name,
            phone=phone,
            schedule_id=schedule_id,
            is_active=True,
        )
        db.add(student)
        db.flush()  # need student.id for package FK
        students_added += 1

        pkg_row = pkg_by_name.get(name)
        if pkg_row is None:
            missing_pkg.append(name)
            continue

        num_classes = pkg_row[3]
        unit_price = pkg_row[4]
        if num_classes is None or unit_price is None:
            missing_pkg.append(name)
            continue

        num_classes = float(num_classes)
        unit_price = float(unit_price)
        # Initial simplification (spec §5): purchased=remaining=total, gifted=0.
        # purchase_price derived from total*unit_price (legacy data; may be 0/negative).
        purchase_price = num_classes * unit_price

        db.add(
            Package(
                student_id=student.id,
                purchased_classes=num_classes,
                gifted_classes=0,
                total_classes=num_classes,
                unit_price=unit_price,
                purchase_price=purchase_price,
                remaining_classes=num_classes,
                start_date=start_date,
                end_date=end_date,
                is_negative=num_classes < 0,
            )
        )
        packages_added += 1

    db.commit()
    print(f"[seed] inserted {students_added} students, {packages_added} packages.")
    if missing_pkg:
        print(f"[seed] WARN: no package matched for {len(missing_pkg)} students -> {missing_pkg}")


def main() -> None:
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        seed_schedules(db)
        seed_users(db)
        seed_students_and_packages(db)
    print("[seed] done.")


if __name__ == "__main__":
    main()
