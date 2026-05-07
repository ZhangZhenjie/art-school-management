"""Excel export builders using openpyxl (spec §4.8 / §八.8).

Layout rules:
  - Chinese column headers
  - Date columns formatted YYYY-MM-DD
  - Numeric columns right-aligned
  - Header row bold + background
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, selectinload

from ..models.class_session import ClassSession
from ..models.revenue_adjustment import RevenueAdjustment
from ..models.schedule import Schedule
from ..models.student import Student
from ..services import students as stu_svc

_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(bold=True)
_RIGHT = Alignment(horizontal="right")


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for idx, v in enumerate(row, start=1):
            if v is None:
                continue
            length = len(str(v))
            widths[idx] = max(widths.get(idx, 0), length)
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(w + 2, 8), 40)


_STATUS_LABEL = {"arrears": "欠费", "low_balance": "余额低", "expiring_soon": "快过期"}


def export_students_packages(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "学生及配套"
    headers = ["学生姓名", "班级", "课程类型", "配套ID", "购买课时", "赠送课时", "剩余课时",
               "课单价", "购买总价", "有效期至", "学生状态"]
    ws.append(headers)
    _style_header(ws, len(headers))

    schedules = {s.id: s for s in db.query(Schedule).all()}
    students = (
        db.query(Student)
        .options(selectinload(Student.packages))
        .filter(Student.is_active.is_(True))
        .order_by(Student.name)
        .all()
    )

    for st in students:
        tags = stu_svc.status_tags(st.packages)
        status_str = "、".join(_STATUS_LABEL.get(t, t) for t in tags) or "正常"
        sched = schedules.get(st.schedule_id) if st.schedule_id else None
        if not st.packages:
            ws.append([
                st.name, st.schedule_id or "", sched.name if sched else "",
                "", "", "", "", "", "", "", status_str,
            ])
            continue
        for p in sorted(st.packages, key=lambda x: (x.start_date, x.id)):
            ws.append([
                st.name,
                st.schedule_id or "",
                sched.name if sched else "",
                p.id,
                p.purchased_classes,
                p.gifted_classes,
                p.remaining_classes,
                p.unit_price,
                p.purchase_price,
                p.end_date.isoformat() if p.end_date else "",
                status_str,
            ])

    for r in range(2, ws.max_row + 1):
        for c in range(4, 10):
            ws.cell(row=r, column=c).alignment = _RIGHT

    _autosize(ws)
    return _to_bytes(wb)


def export_monthly_sessions(db: Session, month: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month} 课时消耗"
    headers = ["学生名", "班级", "课程类型", "应出席", "实际出席", "缺席", "未确认", "消耗课时数"]
    ws.append(headers)
    _style_header(ws, len(headers))

    schedules = {s.id: s for s in db.query(Schedule).all()}
    rows = (
        db.query(ClassSession)
        .filter(ClassSession.session_date.between(f"{month}-01", f"{month}-31"))
        .all()
    )

    by_student: dict[int, dict] = {}
    for r in rows:
        b = by_student.setdefault(r.student_id, {
            "expected": 0, "attended": 0, "absent": 0, "unconfirmed": 0,
            "consumed": 0.0, "schedule_id": r.schedule_id,
        })
        b["expected"] += 1
        if r.confirmed and r.attended:
            b["attended"] += 1
        elif r.confirmed and not r.attended:
            b["absent"] += 1
        else:
            b["unconfirmed"] += 1
        b["consumed"] += r.classes_deducted or 0

    students = {
        s.id: s
        for s in db.query(Student).filter(Student.id.in_(list(by_student.keys()) or [-1])).all()
    }
    for sid, b in sorted(
        by_student.items(),
        key=lambda kv: students[kv[0]].name if kv[0] in students else "",
    ):
        st = students.get(sid)
        if not st:
            continue
        sched = schedules.get(b["schedule_id"]) if b["schedule_id"] else None
        ws.append([
            st.name,
            b["schedule_id"] or "",
            sched.name if sched else "",
            b["expected"],
            b["attended"],
            b["absent"],
            b["unconfirmed"],
            b["consumed"],
        ])

    for r in range(2, ws.max_row + 1):
        for c in range(4, 9):
            ws.cell(row=r, column=c).alignment = _RIGHT
    _autosize(ws)
    return _to_bytes(wb)


def export_revenue(db: Session, month: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month} 营收"
    headers = ["月份", "来源", "学生名", "课程类型", "上课日期", "消耗课时", "课单价", "营收金额", "备注"]
    ws.append(headers)
    _style_header(ws, len(headers))

    schedules = {s.id: s.name for s in db.query(Schedule).all()}
    students = {s.id: s.name for s in db.query(Student).all()}

    sessions = (
        db.query(ClassSession)
        .filter(ClassSession.confirmed.is_(True))
        .filter(ClassSession.revenue_month == month)
        .order_by(ClassSession.session_date, ClassSession.student_id)
        .all()
    )
    for s in sessions:
        if (s.revenue_amount or 0) == 0 and (s.classes_deducted or 0) == 0:
            continue
        unit = (s.revenue_amount / s.classes_deducted) if (s.classes_deducted or 0) else None
        ws.append([
            s.revenue_month or month,
            "课时",
            students.get(s.student_id, "?"),
            schedules.get(s.schedule_id, ""),
            s.session_date.isoformat() if s.session_date else "",
            s.classes_deducted or 0,
            round(unit, 4) if unit is not None else "",
            round(s.revenue_amount or 0, 2),
            "",
        ])

    adjustments = (
        db.query(RevenueAdjustment)
        .filter(RevenueAdjustment.revenue_month == month)
        .order_by(RevenueAdjustment.operated_at)
        .all()
    )
    for a in adjustments:
        ws.append([
            a.revenue_month,
            "调整",
            a.student_name,
            "",
            "",
            a.classes_count,
            a.unit_price,
            round(a.amount, 2),
            f"{a.reason}{' · ' + a.note if a.note else ''}",
        ])

    ws.append([])
    last = ws.max_row + 1
    ws.cell(row=last, column=1, value="合计").font = _HEADER_FONT
    ws.cell(row=last, column=6,
            value=sum((s.classes_deducted or 0) for s in sessions) + sum(a.classes_count for a in adjustments))
    ws.cell(row=last, column=8,
            value=round(sum((s.revenue_amount or 0) for s in sessions) + sum(a.amount for a in adjustments), 2))
    for c in (1, 6, 8):
        ws.cell(row=last, column=c).font = _HEADER_FONT
        ws.cell(row=last, column=c).alignment = _RIGHT

    for r in range(2, last + 1):
        for c in (6, 7, 8):
            ws.cell(row=r, column=c).alignment = _RIGHT

    _autosize(ws)
    return _to_bytes(wb)


def _to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
